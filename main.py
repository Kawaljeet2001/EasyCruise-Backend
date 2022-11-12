from datetime import datetime
from fastapi import FastAPI, Depends, Response
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from database import get_db
from models import FlightModel
from dotenv import load_dotenv
import models
from database import engine
import database
from Schemas import Flight as FlightSchema
from Schemas import Journey as JourneySchema
from Schemas import FlightSchedule as FlightScheduleSchema
from Schemas import User as UserSchema
from utils import seatAllocation
from utils import Flight as FlightUtil
from sqlalchemy.sql import func
database.Base.metadata.create_all(bind=engine)

load_dotenv('.env')
app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173"],
    allow_methods=["*"],
    allow_headers=["*"],
    allow_credentials=True,
)

@app.get("/")
def home():
    res = seatAllocation.performSeatAllocation(100 , 67 , "Economy" , 4)
    return {"hello world" : res}

@app.get("/api/flight")
def getFlight(db : Session = Depends(get_db)):
    # flight = db.query(models.FlightModel).filter(models.FlightModel.id == flightId).first()
    flight = db.query(models.FlightSchedule).all()
    return flight   


@app.post("/api/flight")
def createFlightUtil(data : FlightSchema.FlightCreate, db : Session = Depends(get_db)):
    newFlight = FlightModel(**data.dict())
    db.add(newFlight)
    db.commit()
    db.refresh(newFlight)

    return newFlight

@app.post("/api/scheduleflight/{flightId}")
async def schedule_flight(flightId : int , data : FlightScheduleSchema.FlightScheduleCreate , db : Session = Depends(get_db)):
    new_scheduled_flight = models.FlightSchedule(**data.dict() , flightId=flightId)
    db.add(new_scheduled_flight)
    db.commit()
    db.refresh(new_scheduled_flight)
    return new_scheduled_flight


@app.get("/api/flights/{source}/{destination}")
async def getAvailableflights(source : str , destination: str , db : Session = Depends(get_db)):
    flights = db.query(models.FlightModel).filter(models.FlightModel.source == source and models.FlightModel.destination == destination).all()
    for a in flights:
        a.schedules
    return flights

@app.get("/api/scheduledflights/{source}/{destination}/{cabinClass}/{date}")
async def get_scheduled_flights(source : str , destination: str,cabinClass : str, date : str , db : Session = Depends(get_db)):
    flights = db.query(models.FlightSchedule).join(models.FlightModel).filter(
        models.FlightSchedule.source == source,
        models.FlightSchedule.destination == destination,
        models.FlightSchedule.date == date).order_by(models.FlightModel.busPrice,models.FlightModel.ecoPrice ).all()
    for flight in flights:
        print(flight.flightDetails)
        if cabinClass == "Economy":
            flight.ticketPrice = flight.flightDetails.ecoPrice
        elif cabinClass == "Business":
            flight.ticketPrice = flight.flightDetails.busPrice
        else:
            flight.ticketPrice = flight.flightDetails.execPrice

    return flights


@app.post("/api/user")
async def register_user( data : UserSchema.UserCreate ,db : Session = Depends(get_db)):
    new_user = models.UserModel(**data.dict())
    db.add(new_user)
    db.commit()
    db.refresh(new_user)

    return new_user

@app.post("/api/userlogin")
async def login_user(data : UserSchema.UserBase , response : Response, db : Session = Depends(get_db)):
    user = db.query(models.UserModel).filter(models.UserModel.username == data.username  , models.UserModel.password == data.password).first()

    if(user):
        # send 200 response with cookie set it
        response.set_cookie(key="user" , value=user, httponly=False , samesite="Lax", secure=False, expires=3600)
        return {"message": "The user is successfully logged in" , "data" : user}
    else:
        return {"message": "The user cannot sign in. check username or password" , }


@app.post("/api/userlogout")
async def login_user(response : Response, db : Session = Depends(get_db)):
    response.delete_cookie(key= "user")
    return {"message": "The user is successfully logged out"}


## Journeys
@app.post("/api/journey/{flightId}/{scheduleId}/{userId}/{numberOfTravellers}/{cabinClass}")
async def createJourney(data : JourneySchema.JourneyCreate , flightId : int, scheduleId : int , userId : int, numberOfTravellers : int, cabinClass : str,db : Session = Depends(get_db)):
    new_journey = models.Journey(**data.dict() , flightId=flightId , userId=userId, scheduleId=scheduleId)
    db.add(new_journey)
    db.commit()
    db.refresh(new_journey)

    return new_journey

@app.patch("/api/updateSchedule/{scheduleId}/{cabinClass}/{numTravellers}")
async def update_schedule(scheduleId : int ,cabinClass : str, numTravellers : int , db : Session = Depends(get_db)):
    updated_schedule = db.query(models.FlightSchedule).filter(models.FlightSchedule.id == scheduleId).first()
    print(updated_schedule.flightDetails)
    if cabinClass == "Economy":
        allocated_seats = seatAllocation.performSeatAllocation(updated_schedule.flightDetails.ecoCap , updated_schedule.ecoRem , cabinClass , numTravellers)
    elif cabinClass == "Business":
        allocated_seats = seatAllocation.performSeatAllocation(updated_schedule.flightDetails.busCap , updated_schedule.busRem , cabinClass , numTravellers)
    else:
        allocated_seats = seatAllocation.performSeatAllocation(updated_schedule.flightDetails.execCap , updated_schedule.execRem , cabinClass , numTravellers)


    if cabinClass == "Economy":
        updated_schedule.ecoRem = updated_schedule.ecoRem - numTravellers
    elif cabinClass == "Business":
        updated_schedule.busRem = updated_schedule.busRem - numTravellers
    else:
        updated_schedule.execRem = updated_schedule.execRem - numTravellers
    db.commit()
    db.refresh(updated_schedule)


    return {"schedule" : updated_schedule , "seats" : allocated_seats} 


@app.get("/api/journey/{pnr}")
async def get_journey(pnr : str , db : Session = Depends(get_db)):
    journey = db.query(models.Journey).filter(models.Journey.pnr == pnr).first()
    if journey:
        print(journey.userDetails)
        print(journey.scheduleDetails)
        print(journey.flightDetails)
        return journey
    return journey

@app.get("/api/userjourneys/{userId}")
async def get_user_journeys(userId : int, db : Session = Depends(get_db)):
    journeys = db.query(models.Journey).filter(models.Journey.userId == userId).all()
    for journey in journeys:
        print(journey.scheduleDetails)
        print(journey.flightDetails)
    return journeys

@app.get('/admin/flights')
async def display_admin_flights(db : Session = Depends(get_db)):
    flights = db.query(models.FlightModel).order_by(models.FlightModel.id).all()
    return flights
    # return all the flights in the database

@app.get('/admin/flight/{flightId}')
async def get_admin__flight_details(flightId : int ,db : Session = Depends(get_db)):
    flight = db.query(models.FlightModel).filter(models.FlightModel.id==flightId).first()
    print(flight.schedules)
    return flight
    # return all the flights in the database

@app.get('/admin/flightschedules/{flightId}')
async def get_admin__flight_details(flightId : int ,db : Session = Depends(get_db)):
    flight = db.query(models.FlightSchedule).filter(models.FlightSchedule.flightId==flightId).order_by(models.FlightSchedule.flightId).all()
    # print(flight.schedules)
    return flight
    # return all the flights in the database


@app.get("/admin/overviewStats")
async def get_overview_stats(db : Session = Depends(get_db)):
    # get the number of flights flown from schedules
    flights_flown = db.query(models.FlightSchedule).all()
    aircrafts = db.query(models.FlightModel).all()
    users = db.query(models.UserModel).all()
    totalEarnings = db.query(func.sum(models.Journey.amountPayed)).scalar()

    return {"flights_flown" : flights_flown , "aircrafts" : aircrafts , "users" : users , "earnings" : totalEarnings}
# all flights of a company
# unique flight schedules and details


@app.get("/admin/schedulearnings/{flightId}")
async def get_schedule_earnings(flightId : int, db:Session = Depends(get_db)):
    earnings = db.query(models.Journey.scheduleId, func.sum(models.Journey.amountPayed).label('earnings')).filter(models.Journey.flightId == flightId).order_by(models.Journey.scheduleId).group_by(models.Journey.scheduleId).all()
    return {"earnings" : earnings}


@app.get("/admin/flightearnings/{flightId}")
async def get_total_flight_earnings(flightId : int, db:Session = Depends(get_db)):
    earnings = db.query(models.Journey.flightId, func.sum(models.Journey.amountPayed).label('earnings')).filter(models.Journey.flightId == flightId).group_by(models.Journey.flightId).first()
    return earnings



## controllers for bulk creating data

import openpyxl
@app.post("/bulkflights")
async def bulk_create_flights(db : Session = Depends(get_db)):
    ## load the excel file
    workbook = openpyxl.load_workbook("BulkCreateData.xlsx", data_only=True)
    sheet = workbook['Flight']

    number = 27
    start = 3
    createdFlights = []
    for i in range(start , start + number):
        newFlight = models.FlightModel(
        source=sheet.cell(row=i,column=4).value,
        destination=sheet.cell(row=i,column=5).value,
        company=sheet.cell(row=i,column=3).value,
        ecoCap=sheet.cell(row=i,column=8).value,
        busCap=sheet.cell(row=i,column=9).value,
        execCap=sheet.cell(row=i,column=10).value,
        ecoPrice=sheet.cell(row=i,column=11).value,
        busPrice=sheet.cell(row=i,column=12).value,
        execPrice=sheet.cell(row=i,column=13).value,
        duration=sheet.cell(row=i,column=14).value,
        arrival=sheet.cell(row=i,column=15).value,
        flightCode=sheet.cell(row=i,column=2).value,
        departure=sheet.cell(row=i,column=16).value
        )
        db.add(newFlight)
        db.commit()
        db.refresh(newFlight)

        createdFlights.append(newFlight)
        newFlight = ""
    return createdFlights

